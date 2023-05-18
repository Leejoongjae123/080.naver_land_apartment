import json

import openpyxl
import requests
import pprint
import time
from bs4 import BeautifulSoup
import random
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment




def GetArticleList():
    dataList = []
    saleTypes=['A1','B1','B2']
    # saleTypes = ['B2']
    for saleType in saleTypes:
        for page in range(1,99999):
            cookies = {
                'NNB': 'DSFIMBNTPBPGI',
                'nx_ssl': '2',
                'page_uid': 'ibvVHlp0YiRsseIwJ+0ssssssb8-215902',
                'nid_inf': '-1261052909',
                'NID_AUT': '65nnGO7Y/pMvP7v7KnZYS8Pbn+OIZ5PkpNQ7KwqADJSJrcDEDttAml2c1OIFs2bG',
                'NID_JKL': '2mwLzbbKp4o/E8RoreXlcTNJ5SDT+TsDfaDBT2suDL4=',
                'CBI_SES': 'ngqVPSJS30dXlSa5CHdMm2fTagnvme1hcRkNBOH4FfQIqHqWYna2AXWErLo99+ak3uTHQqvmvn0UCDB/ZXnQmeZhD/3c2ftIJ1NM7STIJCN8oU1Qq+N6EweFXqS65gWQsIPbiCJ3x85TV8PSHNAiEWyGPJuo0JjYKoiR8CeMe1mHGfWJSCfxAygaSUsMqm5bXyqwQTRjFdD2cm4jK9VOwc2mojayCGps/dsYC+C1VON1TIcjWz3XoZtR4G5W7kZhRPKTXYZ8UaSQtuGRLZTpqSqvias35uBDguqTn8NgDW9B3VsGMlwMMyszAMGHTsfNAp3s9qV3IkY71e8iTsxE4inhEuhWVexo6jhY/uDFA/qMoAvhYEOxR42KnoEkY0awSB+hr/OdUn/zOXKHuWmdBwt8kBJpMMw6S3n6BS/RfGTpz2IWVhq13kT94Ocg91NOagRe5cn83msJTtfz8ePxth4onFlELAwfLIjrjVbYQjQ=',
                'CBI_CHK': '"r5V0mf9uRUZHZ/vmLGy3ez7f4/k4aqWXL5o03eN68fqXZTaCNyDz1HDhR9IEUC3fd4oA/UxetbANcLKZ+kov4ppDlaoltyEul9c25Ll+d5CuH6C1ZTqwEY1CMGI/JbHi90xZadZjaip4WHDkcouBWOUGdWOeZluaVfZJxB2aS0M="',
                'SHOW_FIN_BADGE': 'Y',
                'HT': 'HM',
                'NID_SES': 'AAABz7BzK01Bi7FSfYMIrhcEwQdv6q7bBBgVA8K2Uelr/pFVDK+p2wN1flqOYxrisB6KccCN8m/xZYzhs1B0WYgVSF3u5l6IWjVK8RKgqk9l32wvRLZNmNry+ZQQRiPc3ZqJL4x+bfeOAgCfHaGjcfRyx07QCJGa9J8qgrvoJqkwpXw/qUNcpt3qVMKjqa8ZOJoyrw3KGezXG1BeKWwS8y9hFRHLSCJidLwUnD9f/hkbLE8Y5rHIyWnzK/mn0WzmcOc0x8O3l+HwK6U+6V4iDwiqP891GiJugJWgW+v9DOpglCvlFAJb9D2D8P051idXZIyh8i5t/dU9ISHwAz02CDl/G1FoCDLTgt79Chl72YOjrGEvZdS/e9KVFwsMgd6bxBK8ajLBcAIgLTWr2XNZJq86bQuV9pUsLD1zpy1fIB3E9Lu59xTMb+ZMxtsUGNaHLzF3NkOlS/Wo3dMLOCP274Wh5/p2+YgVn7dBxG4dt582+FJWaaGuxyx4/DMh698a2aVg7key8EUMnDddxitkebEvwGsth62mIKP+/HMwZ6pGITF6CYIS4Vwb+Bw3oDoigmL12Kl+c79L8XqgUvPuwsfLAhAdKeKt43tGxfTOo3LVXPDN',
                'wcs_bt': '44058a670db444:1684077813',
                'REALESTATE': 'Mon%20May%2015%202023%2000%3A23%3A33%20GMT%2B0900%20(KST)',
            }

            headers = {
                'Accept': 'application/json, text/javascript, */*; q=0.01',
                'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                'Connection': 'keep-alive',
                # 'Cookie': 'NNB=DSFIMBNTPBPGI; nx_ssl=2; page_uid=ibvVHlp0YiRsseIwJ+0ssssssb8-215902; nid_inf=-1261052909; NID_AUT=65nnGO7Y/pMvP7v7KnZYS8Pbn+OIZ5PkpNQ7KwqADJSJrcDEDttAml2c1OIFs2bG; NID_JKL=2mwLzbbKp4o/E8RoreXlcTNJ5SDT+TsDfaDBT2suDL4=; CBI_SES=ngqVPSJS30dXlSa5CHdMm2fTagnvme1hcRkNBOH4FfQIqHqWYna2AXWErLo99+ak3uTHQqvmvn0UCDB/ZXnQmeZhD/3c2ftIJ1NM7STIJCN8oU1Qq+N6EweFXqS65gWQsIPbiCJ3x85TV8PSHNAiEWyGPJuo0JjYKoiR8CeMe1mHGfWJSCfxAygaSUsMqm5bXyqwQTRjFdD2cm4jK9VOwc2mojayCGps/dsYC+C1VON1TIcjWz3XoZtR4G5W7kZhRPKTXYZ8UaSQtuGRLZTpqSqvias35uBDguqTn8NgDW9B3VsGMlwMMyszAMGHTsfNAp3s9qV3IkY71e8iTsxE4inhEuhWVexo6jhY/uDFA/qMoAvhYEOxR42KnoEkY0awSB+hr/OdUn/zOXKHuWmdBwt8kBJpMMw6S3n6BS/RfGTpz2IWVhq13kT94Ocg91NOagRe5cn83msJTtfz8ePxth4onFlELAwfLIjrjVbYQjQ=; CBI_CHK="r5V0mf9uRUZHZ/vmLGy3ez7f4/k4aqWXL5o03eN68fqXZTaCNyDz1HDhR9IEUC3fd4oA/UxetbANcLKZ+kov4ppDlaoltyEul9c25Ll+d5CuH6C1ZTqwEY1CMGI/JbHi90xZadZjaip4WHDkcouBWOUGdWOeZluaVfZJxB2aS0M="; SHOW_FIN_BADGE=Y; HT=HM; NID_SES=AAABz7BzK01Bi7FSfYMIrhcEwQdv6q7bBBgVA8K2Uelr/pFVDK+p2wN1flqOYxrisB6KccCN8m/xZYzhs1B0WYgVSF3u5l6IWjVK8RKgqk9l32wvRLZNmNry+ZQQRiPc3ZqJL4x+bfeOAgCfHaGjcfRyx07QCJGa9J8qgrvoJqkwpXw/qUNcpt3qVMKjqa8ZOJoyrw3KGezXG1BeKWwS8y9hFRHLSCJidLwUnD9f/hkbLE8Y5rHIyWnzK/mn0WzmcOc0x8O3l+HwK6U+6V4iDwiqP891GiJugJWgW+v9DOpglCvlFAJb9D2D8P051idXZIyh8i5t/dU9ISHwAz02CDl/G1FoCDLTgt79Chl72YOjrGEvZdS/e9KVFwsMgd6bxBK8ajLBcAIgLTWr2XNZJq86bQuV9pUsLD1zpy1fIB3E9Lu59xTMb+ZMxtsUGNaHLzF3NkOlS/Wo3dMLOCP274Wh5/p2+YgVn7dBxG4dt582+FJWaaGuxyx4/DMh698a2aVg7key8EUMnDddxitkebEvwGsth62mIKP+/HMwZ6pGITF6CYIS4Vwb+Bw3oDoigmL12Kl+c79L8XqgUvPuwsfLAhAdKeKt43tGxfTOo3LVXPDN; wcs_bt=44058a670db444:1684077813; REALESTATE=Mon%20May%2015%202023%2000%3A23%3A33%20GMT%2B0900%20(KST)',
                'Referer': 'https://m.land.naver.com/',
                'Sec-Fetch-Dest': 'empty',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Site': 'same-origin',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
                'X-Requested-With': 'XMLHttpRequest',
                'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
            }

            params = {
                'hscpNo': '1808',
                'cortarNo': '4111710500',
                'tradTpCd': str(saleType),
                'order': 'point_',
                'showR0': 'N',
                'page': str(page),
            }

            response = requests.get('https://m.land.naver.com/complex/getComplexArticleList', params=params, cookies=cookies, headers=headers)


            results=json.loads(response.text)['result']['list']

            print("매물수:",len(results))
            # pprint.pprint(results)
            # pprint.pprint(resultsRaw)

            if len(results)==0:
                break
            try:
                title = results[0]['atclNm']
            except:
                title = ""
            # print("단지명:", title)
            for result in results:
                # print(result['atclNo'])
                dataList.append(result['atclNo'])
            time.sleep(1)
    return dataList, title
def GetDetailInfo(dataElem):
    cookies = {
        'NNB': 'DSFIMBNTPBPGI',
        'nx_ssl': '2',
        'page_uid': 'ibvVHlp0YiRsseIwJ+0ssssssb8-215902',
        'SHOW_FIN_BADGE': 'Y',
        'HT': 'HM',
        'nid_inf': '-1221207775',
        'NID_AUT': 'EKkPd7vrzwWHJ+vh9kh8/3nsc13W9oJaNHpCur5VqjsmR0FjEa/pnjZPx5NtN0e7',
        'NID_SES': 'AAABqIHQy5/NnM2DS1C/wriE4UK52EL5psgGl/gp3SPQ46ZFoq252wajBbSKPPcHUXTYS//TwPNa2FA2e2gXa0OGSYWsFiyv/s+HAHMD9B9+38/Msto8xskULHJL7TOZChfd/Odh5tAXl4sWUXngF0fEoq+Lf0h52kFxvA+AJLVw1ZGOhFcvgtuqOejYrkMCZfILNxHsRFOat9IW5tJk0XtdbEmSAx7uJROIOcRHcSCu2ZwLeNHMRFQuvVwhcAiHfusbrzE5NfIsI7Fa4qO7WcxvZtREVQ9vhiKbvdST51qOTlQzcwivfsMUjVCXq9f7k5YevPeIonSPdQGqrp8pZt4cNfNYICGiQGJga8R/JUMu8tJNeQJG27M7mfQneJNXZ8/HtlTlDzltQoTTTk8M+SJrLPgWJmrffQRmP22psBCSbtxu+HYMSOE4oDmp+mUy/k6CAR91uSieTBItM+6oc2koqH0lucws4cVnnD/27ZPLloljN5Q80JPkNoBw5jrfj7jzoqeXLFRMdcnKk0O/UO+nFG1kbzSGf+6RZClgG5AI0/4zZpFK9SJslbQ2irgQEslbNw==',
        'NID_JKL': 'crmO14b07iPM3sARocsufi/49DTuiPZW9qf0A2nLUrM=',
        'CBI_SES': '4lwwURCQtV8BAh0OIPxGBztqvkDLwr+OhM913yqLG9MKVE+zQaZgsvIRuUwcsDjZF0yzfUu1piYlwQzq6y2neVm7sZYsvsW04EiNUp7iK1Lgm+31fnOL5hqSrxOMUNsYfkpJ3jTj0gDoXWiXTs4f/tciG/92pM4TSCTikzVqWJk5SCpXm+Y9ScafV8EBnUJTHaKH8B10C2u5Fm1/l/lzF6t6CkOjIgLLUuoacXFlbtsQJlseqQQlNHEUI/eDBIdOeBbiqIuZvzk4qNlMFUIdudImRCAtY4T3RPIsWB5W/dhRxUiZbP3oXh2cPrLkp9w632CbvcC09RJz8jTekO+TzDAXj0Un5Hcf+WqzqXG04GcJmDPbG1PgzY4Ao4+JjqrecYzyMAuFQlxaklkvX+BIvqjfIT+VdgoYj4D82QiYWO4FDVip8DvCMrVglUGypbl5jujbA8jqxYIJnxsrwhIlbQ7ncNilEPsZQXYqtERQ9UY=',
        'CBI_CHK': '"r5V0mf9uRUZHZ/vmLGy3ez7f4/k4aqWXL5o03eN68foBDFj5tVHn3+kHsyl+z6JdR3r+R9G2kW4ZqiHW74Bt3jUByVuTyi/ZCd7cE+nndRAiby6PFkTuqcrLT8UudhvaaH0PH6+m/lGjV/nTwYPhLitPBltJjmtNG3USAm1Prdw="',
        'wcs_bt': '44058a670db444:1684078649',
        'REALESTATE': '1684078655224',
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        # 'Cookie': 'NNB=DSFIMBNTPBPGI; nx_ssl=2; page_uid=ibvVHlp0YiRsseIwJ+0ssssssb8-215902; SHOW_FIN_BADGE=Y; HT=HM; nid_inf=-1221207775; NID_AUT=EKkPd7vrzwWHJ+vh9kh8/3nsc13W9oJaNHpCur5VqjsmR0FjEa/pnjZPx5NtN0e7; NID_SES=AAABqIHQy5/NnM2DS1C/wriE4UK52EL5psgGl/gp3SPQ46ZFoq252wajBbSKPPcHUXTYS//TwPNa2FA2e2gXa0OGSYWsFiyv/s+HAHMD9B9+38/Msto8xskULHJL7TOZChfd/Odh5tAXl4sWUXngF0fEoq+Lf0h52kFxvA+AJLVw1ZGOhFcvgtuqOejYrkMCZfILNxHsRFOat9IW5tJk0XtdbEmSAx7uJROIOcRHcSCu2ZwLeNHMRFQuvVwhcAiHfusbrzE5NfIsI7Fa4qO7WcxvZtREVQ9vhiKbvdST51qOTlQzcwivfsMUjVCXq9f7k5YevPeIonSPdQGqrp8pZt4cNfNYICGiQGJga8R/JUMu8tJNeQJG27M7mfQneJNXZ8/HtlTlDzltQoTTTk8M+SJrLPgWJmrffQRmP22psBCSbtxu+HYMSOE4oDmp+mUy/k6CAR91uSieTBItM+6oc2koqH0lucws4cVnnD/27ZPLloljN5Q80JPkNoBw5jrfj7jzoqeXLFRMdcnKk0O/UO+nFG1kbzSGf+6RZClgG5AI0/4zZpFK9SJslbQ2irgQEslbNw==; NID_JKL=crmO14b07iPM3sARocsufi/49DTuiPZW9qf0A2nLUrM=; CBI_SES=4lwwURCQtV8BAh0OIPxGBztqvkDLwr+OhM913yqLG9MKVE+zQaZgsvIRuUwcsDjZF0yzfUu1piYlwQzq6y2neVm7sZYsvsW04EiNUp7iK1Lgm+31fnOL5hqSrxOMUNsYfkpJ3jTj0gDoXWiXTs4f/tciG/92pM4TSCTikzVqWJk5SCpXm+Y9ScafV8EBnUJTHaKH8B10C2u5Fm1/l/lzF6t6CkOjIgLLUuoacXFlbtsQJlseqQQlNHEUI/eDBIdOeBbiqIuZvzk4qNlMFUIdudImRCAtY4T3RPIsWB5W/dhRxUiZbP3oXh2cPrLkp9w632CbvcC09RJz8jTekO+TzDAXj0Un5Hcf+WqzqXG04GcJmDPbG1PgzY4Ao4+JjqrecYzyMAuFQlxaklkvX+BIvqjfIT+VdgoYj4D82QiYWO4FDVip8DvCMrVglUGypbl5jujbA8jqxYIJnxsrwhIlbQ7ncNilEPsZQXYqtERQ9UY=; CBI_CHK="r5V0mf9uRUZHZ/vmLGy3ez7f4/k4aqWXL5o03eN68foBDFj5tVHn3+kHsyl+z6JdR3r+R9G2kW4ZqiHW74Bt3jUByVuTyi/ZCd7cE+nndRAiby6PFkTuqcrLT8UudhvaaH0PH6+m/lGjV/nTwYPhLitPBltJjmtNG3USAm1Prdw="; wcs_bt=44058a670db444:1684078649; REALESTATE=1684078655224',
        'Referer': 'https://m.land.naver.com/',
        'Sec-Fetch-Dest': 'iframe',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    response = requests.get('https://m.land.naver.com/article/info/{}?newMobile'.format(dataElem), cookies=cookies, headers=headers)
    soup=BeautifulSoup(response.text,'lxml')
    scripts=soup.find_all('script')
    targetScript=""
    for script in scripts:
        if str(script).find("roomNumber")>=0:
            targetScript=script
    positionFr=str(targetScript).find("{")
    positionRr=str(targetScript).rfind("}")
    result=json.loads(str(targetScript)[positionFr:positionRr+1])['state']
    # pprint.pprint(result)
    # with open('output.json', 'w') as f:
    # 	json.dump(result, f, indent=2)
    try:
        articleNo=result['article']['addition']['articleNo']
    except:
        articleNo=""
    # print('articleNo:',articleNo)
    try:
        buildingName=result['article']['addition']['buildingName'].replace("동","")
    except:
        buildingName=""
    # print('buildingName:',buildingName)
    try:
        roomNumber=result['article']['article']['roomNumber']
    except:
        roomNumber=""
    # print('roomNumber:',roomNumber)
    try:
        tradeType=result['article']['addition']['tradeTypeName']
    except:
        tradeType =""
    # print('tradeType:',tradeType)
    try:
        price=result['article']['addition']['dealOrWarrantPrc']
    except:
        price =""
    # print('price:',price)
    try:
        exclusiveSpace=result['article']['space']['exclusiveSpace']
    except:
        exclusiveSpace =""
    # print('exclusiveSpace:',exclusiveSpace)
    try:
        supplySpace=result['article']['space']['supplySpace']
    except:
        supplySpace =""
    # print('supplySpace:',supplySpace)
    try:
        directionTypeName=result['article']['facility']['directionTypeName']
    except:
        directionTypeName =""
    # print('directionTypeName:',directionTypeName)
    try:
        directionBaseTypeName=result['article']['facility']['directionBaseTypeName']
    except:
        directionBaseTypeName =""
    # print('directionBaseTypeName:',directionBaseTypeName)
    try:
        correspondingFloorCount=result['article']['floor']['correspondingFloorCount']
    except:
        correspondingFloorCount =""
    # print('correspondingFloorCount:',correspondingFloorCount)
    try:
        totalFloorCount=result['article']['floor']['totalFloorCount']
    except:
        totalFloorCount =""
    # print('totalFloorCount:',totalFloorCount)
    try:
        roomCount=result['article']['article']['roomCount']
    except:
        roomCount=""
    # print('roomCount:',roomCount)
    try:
        bathroomCount=result['article']['article']['bathroomCount']
    except:
        bathroomCount=""
    # print('bathroomCount:',bathroomCount)
    try:
        householdCountByPtp=result['article']['article']['householdCountByPtp']
    except:
        householdCountByPtp=""
    # print('householdCountByPtp:',householdCountByPtp)
    try:
        aptHouseholdCount=result['article']['article']['aptHouseholdCount']
    except:
        aptHouseholdCount=""
    # print('aptHouseholdCount:',aptHouseholdCount)
    try:
        moveInTypeName=result['article']['article']['moveInTypeName']
    except:
        moveInTypeName=""
    # print('moveInTypeName:',moveInTypeName)
    try:
        detailDescription=result['article']['article']['detailDescription']
    except:
        detailDescription=""
    # print('detailDescription:',detailDescription)
    try:
        articleConfirmYmd=result['article']['addition']['articleConfirmYmd']
    except:
        articleConfirmYmd=""
    # print('articleConfirmYmd:',articleConfirmYmd)
    try:
        rentPrc=result['article']['addition']['rentPrc']
    except:
        rentPrc=""
    if len(rentPrc)>=1:
        price=price+"/"+rentPrc
    data=[articleNo,buildingName,roomNumber,tradeType,price,"{}㎡[{}㎡]".format(exclusiveSpace,supplySpace),'{}[{}]'.format(directionTypeName,directionBaseTypeName),"{}[{}]".format(correspondingFloorCount,totalFloorCount),"방{}/욕실{}".format(roomCount,bathroomCount),"{}[총{}세대]".format(householdCountByPtp,aptHouseholdCount),moveInTypeName,detailDescription,articleConfirmYmd]
    print(data)
    return data
    print("================================")


import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *


class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(int)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.

    def run(self):
        dataList, title = GetArticleList()
        wb = openpyxl.Workbook()
        ws = wb.active
        columnName = ['매물번호', '동', '호수', '분류', '가격', '전용면적[공급면적]', '방향', '해당층[총층]', '방수/욕실수', '해당면적세대수[총세대수]', '입주가능일',
                      '매물설명', '등록일']
        ws.append(columnName)
        timeNowString = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        for dataElem in dataList:
            timeNow = datetime.datetime.now()
            print("현재시간:", timeNow)
            data = GetDetailInfo(dataElem)
            ws.append(data)
            wb.save('{}_{}.xlsx'.format(title, timeNowString))
            time.sleep(random.randint(10, 15) * 0.1)

        noRow = ws.max_row
        for row in range(1, noRow + 1):
            for col in range(1, 15):
                if col == 6 or col == 7 or col == 10:
                    ws.column_dimensions[get_column_letter(col)].width = 15
                elif col == 12:
                    ws.column_dimensions[get_column_letter(col)].width = 20
                else:
                    ws.column_dimensions[get_column_letter(col)].width = 12

                ws["{}{}".format(get_column_letter(col), row)].alignment = Alignment(horizontal='center')
                print(get_column_letter(col))
        wb.save('{}_{}.xlsx'.format(title, timeNowString))
    def stop(self):
        pass

class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        QApplication.processEvents()

    def start(self):
        print('11')
        self.x = Thread(self)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))

    def setSlot(self):
        pass

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())




